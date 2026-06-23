# -*- coding: utf-8 -*-
"""Конструктор отчётов: произвольный разрез Директа за свободный период.

Те же запросы Reports API, что и недельная рассылка, но гибко:
  * базовый разрез (уровень): аккаунт/кампании/группы/объявления/фразы/поисковые запросы;
  * дополнительные измерения (срезы): дата (день/неделя/месяц), устройство, пол, возраст,
    гео, тип сети — можно несколько сразу;
  * модель атрибуции и свободные даты.

Конверсии: поле Conversions Reports API отдаёт ДАЖЕ без целей и без Метрики — поэтому
конверсии/CR/CPA показываются для любого клиента. Если у клиента заданы цели, конверсии
считаются по ним под выбранной моделью атрибуции (точнее), иначе берётся «голое» Conversions.
"""
from datetime import date as _date, timedelta as _td

from . import report as R

# базовый уровень -> (тип отчёта Директа, поля-измерения сущности, заголовки)
LEVELS = {
    "account":     ("ACCOUNT_PERFORMANCE_REPORT", [], []),
    "campaign":    ("CAMPAIGN_PERFORMANCE_REPORT", ["CampaignName"], ["Кампания"]),
    "adgroup":     ("ADGROUP_PERFORMANCE_REPORT", ["CampaignName", "AdGroupName"], ["Кампания", "Группа"]),
    "ad":          ("AD_PERFORMANCE_REPORT", ["CampaignName", "AdGroupName", "AdId"], ["Кампания", "Группа", "ID объявл."]),
    "keyword":     ("CRITERIA_PERFORMANCE_REPORT", ["CampaignName", "AdGroupName", "Criterion"], ["Кампания", "Группа", "Фраза"]),
    "searchquery": ("SEARCH_QUERY_PERFORMANCE_REPORT", ["CampaignName", "Query"], ["Кампания", "Запрос"]),
}
LEVEL_LABELS = {
    "account": "Аккаунт (сводка)", "campaign": "Кампании", "adgroup": "Группы объявлений",
    "ad": "Объявления", "keyword": "Фразы", "searchquery": "Поисковые запросы",
}

# дополнительные измерения (срезы): ключ -> (FieldName Директа, заголовок). Порядок важен.
SEGMENTS = {
    "date":    ("Date", "Дата"),
    "device":  ("Device", "Устройство"),
    "gender":  ("Gender", "Пол"),
    "age":     ("Age", "Возраст"),
    "geo":     ("LocationOfPresenceName", "Гео"),
    "network": ("AdNetworkType", "Сеть"),
}
SEGMENT_ORDER = ["date", "device", "gender", "age", "geo", "network"]
DATE_GRAINS = {"day": "По дням", "week": "По неделям", "month": "По месяцам"}

# модели атрибуции Директа (максимум): обычные + кросс-девайс (…D) + авто
ATTRIBUTION_MODELS = ["LSC", "LC", "FC", "LYDC", "LSCD", "LCD", "FCD", "LYDCD", "AUTO"]
CONV_LEVELS = {"account", "campaign", "adgroup", "ad", "keyword"}   # где доступны конверсии

# колонки-метрики, которые можно выбрать к показу
METRICS = [
    {"key": "cost", "label": "Расход"}, {"key": "imp", "label": "Показы"}, {"key": "clicks", "label": "Клики"},
    {"key": "ctr", "label": "CTR"}, {"key": "cpc", "label": "CPC"},
    {"key": "conv", "label": "Конверсии"}, {"key": "cr", "label": "CR"}, {"key": "cpa", "label": "CPA"},
]


def options():
    return {
        "levels": [{"key": k, "label": LEVEL_LABELS[k]} for k in LEVELS],
        "attributions": ATTRIBUTION_MODELS,
        "conv_levels": sorted(CONV_LEVELS),
        "segments": [{"key": k, "label": SEGMENTS[k][1]} for k in SEGMENT_ORDER],
        "date_grains": [{"key": k, "label": DATE_GRAINS[k]} for k in ("day", "week", "month")],
        "metrics": METRICS,
    }


def _date_bucket(v, grain):
    """Свернуть дату YYYY-MM-DD к дню/неделе/месяцу для группировки на нашей стороне."""
    v = (v or "").strip()
    if grain == "day" or len(v) < 10:
        return v
    try:
        dt = _date(int(v[0:4]), int(v[5:7]), int(v[8:10]))
    except ValueError:
        return v
    if grain == "month":
        return v[0:7]
    if grain == "week":
        mon = dt - _td(days=dt.weekday())
        return "нед. с " + mon.isoformat()
    return v


def _metrics(imp, clk, cost, conv):
    return {
        "imp": imp, "clicks": clk, "cost": cost, "conv": conv,
        "ctr": (clk / imp * 100 if imp else 0),
        "cpc": (cost / clk if clk else 0),
        "cr": (conv / clk * 100 if clk else 0),
        "cpa": (cost / conv if conv else 0),
    }


def build(token, login, level, date_from, date_to, attribution="LSC", goal_defs=None,
          segments=None, date_grain="day", limit=100, _post=None, _sleep=None):
    if level not in LEVELS:
        raise RuntimeError("Неизвестный разрез: {}".format(level))
    rtype, ent_dims, ent_titles = LEVELS[level]
    segs = [s for s in SEGMENT_ORDER if s in (segments or [])]
    seg_fields = [SEGMENTS[s][0] for s in segs]
    seg_titles = [SEGMENTS[s][1] for s in segs]
    dim_fields = list(ent_dims) + seg_fields
    dim_titles = list(ent_titles) + seg_titles

    goal_defs = goal_defs or []
    conv_capable = level in CONV_LEVELS
    use_goals = bool(goal_defs) and conv_capable
    goal_ids = [g["id"] for g in goal_defs] if use_goals else None
    attribution = attribution if attribution in ATTRIBUTION_MODELS else "LSC"
    has_date = "date" in segs

    fields = list(dim_fields) + ["Impressions", "Clicks", "Cost"]
    if conv_capable:
        fields.append("Conversions")   # работает даже без Goals

    raw = R.fetch_report(token, login, date_from, date_to, fields,
                         goal_ids=goal_ids, attribution=attribution, report_type=rtype,
                         _post=_post, _sleep=_sleep)

    # агрегируем на своей стороне по кортежу значений измерений (нужно для свёртки дат
    # день->неделя/месяц и чтобы суммы сходились).
    agg = {}
    for r in raw:
        disp = []
        for d in dim_fields:
            val = str(r.get(d) or "")
            if d == "Date":
                val = _date_bucket(val, date_grain)
            disp.append(val)
        imp = R.parse_num(r.get("Impressions"))
        clk = R.parse_num(r.get("Clicks"))
        cost = R.parse_num(r.get("Cost"))
        if use_goals:
            conv = 0.0
            for g in goal_defs:
                col = R._find_goal_col(r, g["id"])
                if col:
                    conv += R.parse_num(r.get(col))
        elif conv_capable:
            conv = R.parse_num(r.get("Conversions"))
        else:
            conv = 0.0
        key = tuple(disp)
        a = agg.get(key)
        if a is None:
            agg[key] = [disp, imp, clk, cost, conv]
        else:
            a[1] += imp; a[2] += clk; a[3] += cost; a[4] += conv

    if not dim_fields:
        # уровень «аккаунт» без срезов — одна строка-сводка
        if agg:
            v = next(iter(agg.values()))
            v[0] = ["Весь аккаунт"]
        dim_titles = ["Аккаунт"]

    rows = [{"dims": (v[0] or ["—"]), "imp": v[1], "clk": v[2], "cost": v[3], "conv": v[4]}
            for v in agg.values()]
    t_imp = sum(x["imp"] for x in rows)
    t_clk = sum(x["clk"] for x in rows)
    t_cost = sum(x["cost"] for x in rows)
    t_conv = sum(x["conv"] for x in rows)

    rows.sort(key=lambda x: -x["cost"])
    n_total = len(rows)
    rows = rows[:max(1, int(limit or 100))]
    out_rows = [{"dims": r["dims"], "m": _metrics(r["imp"], r["clk"], r["cost"], r["conv"])} for r in rows]

    return {
        "level": level, "level_label": LEVEL_LABELS[level],
        "dim_titles": dim_titles or ["Аккаунт"],
        "attribution": attribution, "use_conv": conv_capable, "by_goals": use_goals,
        "segments": segs, "date_grain": date_grain if has_date else None,
        "date_from": date_from, "date_to": date_to,
        "rows": out_rows, "totals": _metrics(t_imp, t_clk, t_cost, t_conv),
        "n_total": n_total, "n_shown": len(out_rows),
    }


def to_text(login, client_name, res, top=25):
    """Текст для Telegram/копирования (топ-N строк по расходу)."""
    seg_note = ""
    if res.get("segments"):
        seg_note = " · срезы: " + ", ".join(SEGMENTS[s][1] for s in res["segments"])
    L = [
        "Отчёт: {} ({})".format(client_name or login, login),
        "Разрез: {} · Атрибуция: {} · Период: {} — {}{}".format(
            res["level_label"], res["attribution"], res["date_from"], res["date_to"], seg_note),
        "",
        "ИТОГО:",
        "  Расход: {}  Показы: {}  Клики: {}".format(
            R.fmt_money(res["totals"]["cost"]), R.fmt_int(res["totals"]["imp"]), R.fmt_int(res["totals"]["clicks"])),
        "  CTR: {}  CPC: {}".format(R.fmt_pct(res["totals"]["ctr"]), R.fmt_money(res["totals"]["cpc"])),
    ]
    if res["use_conv"]:
        L.append("  Конверсии: {}  CR: {}  CPA: {}".format(
            R.fmt_int(res["totals"]["conv"]), R.fmt_pct(res["totals"]["cr"]), R.fmt_money(res["totals"]["cpa"])))
    multi = res["level"] != "account" or res.get("segments")
    if multi:
        L.append("")
        L.append("Топ {} по расходу:".format(min(top, res["n_shown"])))
        for i, row in enumerate(res["rows"][:top], 1):
            name = " / ".join([d for d in row["dims"] if d]) or "—"
            m = row["m"]
            line = "{}) {} — {}, кликов {}".format(i, name, R.fmt_money(m["cost"]), R.fmt_int(m["clicks"]))
            if res["use_conv"]:
                line += ", конв. {}".format(R.fmt_int(m["conv"]))
            L.append(line)
        if res["n_total"] > res["n_shown"]:
            L.append("… показано {} из {} строк.".format(res["n_shown"], res["n_total"]))
    return "\n".join(L)


def to_xlsx(res, path):
    """Выгружает результат в .xlsx (нужен openpyxl). Колонки: измерения + метрики."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    headers = list(res["dim_titles"]) + ["Расход", "Показы", "Клики", "CTR %", "CPC"]
    if res["use_conv"]:
        headers += ["Конверсии", "CR %", "CPA"]
    ws.append(headers)

    def cells(m):
        out = [round(m["cost"], 2), int(round(m["imp"])), int(round(m["clicks"])),
               round(m["ctr"], 2), round(m["cpc"], 2)]
        if res["use_conv"]:
            out += [int(round(m["conv"])), round(m["cr"], 2), round(m["cpa"], 2)]
        return out

    ndim = len(res["dim_titles"])
    ws.append(["ИТОГО"] + [""] * (ndim - 1) + cells(res["totals"]))
    for row in res["rows"]:
        dims = list(row["dims"]) + [""] * (ndim - len(row["dims"]))
        ws.append(dims + cells(row["m"]))

    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(48, len(str(h)) + 4))
    try:
        ws.freeze_panes = "A2"
    except Exception:  # noqa: BLE001
        pass
    wb.save(path)
    return path
